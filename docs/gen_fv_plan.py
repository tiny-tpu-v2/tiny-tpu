"""
gen_fv_plan.py
Generates docs/tiny_tpu_FV_Plan.xlsx from the verified assertion catalog.
Run with the project venv Python.
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
import os

# ── colour palette ──────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1F3864")   # dark navy  – header
SUB_FILL   = PatternFill("solid", fgColor="2E75B6")   # mid blue   – section header
ALT_FILL   = PatternFill("solid", fgColor="D9E1F2")   # light blue – alternating row
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
# status colours
OPEN_FILL  = PatternFill("solid", fgColor="FFF2CC")   # yellow
PASS_FILL  = PatternFill("solid", fgColor="E2EFDA")   # green
FAIL_FILL  = PatternFill("solid", fgColor="FCE4D6")   # red

HDR_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
SUB_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
BODY_FONT  = Font(name="Calibri", size=10)
MONO_FONT  = Font(name="Consolas", size=9)

THIN  = Side(style="thin",   color="B8CCE4")
MED   = Side(style="medium", color="2E75B6")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MED_BORDER  = Border(left=MED,  right=MED,  top=MED,  bottom=MED)

WRAP  = Alignment(wrap_text=True, vertical="top")
CTR   = Alignment(wrap_text=True, vertical="center", horizontal="center")


def hdr_cell(ws, row, col, value, merge_end=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill, c.font, c.alignment, c.border = HDR_FILL, HDR_FONT, CTR, THIN_BORDER
    if merge_end:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=merge_end)
    return c


def sub_cell(ws, row, col, value, merge_end=None):
    c = ws.cell(row=row, column=col, value=value)
    c.fill, c.font, c.alignment, c.border = SUB_FILL, SUB_FONT, CTR, THIN_BORDER
    if merge_end:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=merge_end)
    return c


def body_cell(ws, row, col, value, alt=False, mono=False, center=False):
    c = ws.cell(row=row, column=col, value=value)
    c.fill   = ALT_FILL if alt else WHITE_FILL
    c.font   = MONO_FONT if mono else BODY_FONT
    c.alignment = CTR if center else WRAP
    c.border = THIN_BORDER
    return c


def status_cell(ws, row, col, value):
    c = ws.cell(row=row, column=col, value=value)
    c.font   = Font(name="Calibri", bold=True, size=10)
    c.alignment = CTR
    c.border = THIN_BORDER
    if value == "Open":
        c.fill = OPEN_FILL
    elif value in ("Pass", "Proven"):
        c.fill = PASS_FILL
    elif value in ("Fail", "CEX"):
        c.fill = FAIL_FILL
    else:
        c.fill = WHITE_FILL
    return c


# ════════════════════════════════════════════════════════════════════════════
#  DATA  – master assertion catalog (all 10 modules)
# ════════════════════════════════════════════════════════════════════════════
# Columns: ID | Module | SVA Name | Category | Priority | Property | ProofType | Bound | Status

ASSERTIONS = [
    # ── pe ───────────────────────────────────────────────────────────────────
    ("PE-A01","pe","p_rst_clears_psum","RST","P1",
     "rst |=> pe_psum_out==0","Unbounded","–","Open"),
    ("PE-A02","pe","p_rst_clears_valid","RST","P1",
     "rst |=> !pe_valid_out","Unbounded","–","Open"),
    ("PE-A03","pe","p_rst_clears_switch","RST","P1",
     "rst |=> !pe_switch_out","Unbounded","–","Open"),
    ("PE-A04","pe","p_rst_clears_weight_out","RST","P1",
     "rst |=> pe_weight_out==0","Unbounded","–","Open"),
    ("PE-A05","pe","p_rst_clears_input_out","RST","P1",
     "rst |=> pe_input_out==0","Unbounded","–","Open"),
    ("PE-A06a","pe","p_disabled_clears_psum","RST","P1",
     "!pe_enabled |=> pe_psum_out==0","Unbounded","–","Open"),
    ("PE-A06b","pe","p_disabled_clears_valid","RST","P1",
     "!pe_enabled |=> !pe_valid_out","Unbounded","–","Open"),
    ("PE-A06c","pe","p_disabled_clears_switch_out","RST","P1",
     "!pe_enabled |=> !pe_switch_out","Unbounded","–","Open"),
    ("PE-A06d","pe","p_disabled_clears_weight_out","RST","P1",
     "!pe_enabled |=> pe_weight_out==0","Unbounded","–","Open"),
    ("PE-A07","pe","p_valid_out_registered","VP","P1",
     "1'b1 |=> pe_valid_out==$past(pe_valid_in)","BMC","4","Open"),
    ("PE-A08","pe","p_switch_out_registered","VP","P1",
     "1'b1 |=> pe_switch_out==$past(pe_switch_in)","BMC","4","Open"),
    ("PE-A09","pe","p_weight_out_when_accepting","DP","P1",
     "pe_accept_w_in |=> pe_weight_out==$past(pe_weight_in)","BMC","4","Open"),
    ("PE-A10","pe","p_weight_out_zero_when_not_accepting","DP","P1",
     "!pe_accept_w_in |=> pe_weight_out==0","BMC","4","Open"),
    ("PE-A11","pe","p_input_out_captured_on_valid","DP","P2",
     "pe_valid_in |=> pe_input_out==$past(pe_input_in)","BMC","4","Open"),
    ("PE-A12","pe","p_psum_zero_when_invalid","DP","P1",
     "!pe_valid_in |=> pe_psum_out==0","BMC","4","Open"),
    ("PE-A13","pe","p_valid_out_low_when_in_low","VP","P1",
     "!pe_valid_in |=> !pe_valid_out","BMC","4","Open"),
    ("PE-A14a","pe","p_rst_clears_weight_reg_active","RST","P1",
     "(rst||!pe_enabled) |=> weight_reg_active==0","Unbounded","–","Open"),
    ("PE-A14b","pe","p_rst_clears_weight_reg_inactive","RST","P1",
     "(rst||!pe_enabled) |=> weight_reg_inactive==0","Unbounded","–","Open"),
    ("PE-A15","pe","p_weight_switch","DP","P1",
     "pe_switch_in |=> weight_reg_active==$past(weight_reg_inactive)","BMC","4","Open"),
    ("PE-A16","pe","p_input_out_clear_when_invalid","DP","P1",
     "!pe_valid_in |=> pe_input_out==16'b0","BMC","4","Open"),
    ("PE-A17","pe","p_rst_clears_overflow","RST","P1",
     "(rst||!pe_enabled) |=> !pe_overflow_out","Unbounded","–","Open"),
    ("PE-A18","pe","p_overflow_is_sticky","FA","P1",
     "pe_overflow_out |=> pe_overflow_out","BMC","4","Open"),
    ("PE-A19","pe","p_mac_zero_input_passthrough_psum","FA","P2",
     "(pe_valid_in && pe_input_in==0) |=> pe_psum_out==$past(pe_psum_in)  [port-observable MAC proxy; see PE-W01]","BMC","4","Open"),

    # ── systolic ─────────────────────────────────────────────────────────────
    ("SYS-A01","systolic","p_rst_clears_valid_out_21","RST","P1",
     "rst |=> !sys_valid_out_21","Unbounded","–","Open"),
    ("SYS-A02","systolic","p_rst_clears_valid_out_22","RST","P1",
     "rst |=> !sys_valid_out_22","Unbounded","–","Open"),
    ("SYS-A03","systolic","p_rst_clears_data_out_21","RST","P1",
     "rst |=> sys_data_out_21==0","Unbounded","–","Open"),
    ("SYS-A04","systolic","p_rst_clears_data_out_22","RST","P1",
     "rst |=> sys_data_out_22==0","Unbounded","–","Open"),
    ("SYS-A05","systolic","p_valid_21_one_cycle_delay","VP","P1",
     "sys_start_2 |=> sys_valid_out_21","BMC","8","Open"),
    ("SYS-A06","systolic","p_valid_22_three_cycles_after_start1","VP","P1",
     "sys_start_1 |=> ##2 sys_valid_out_22  [3 register stages: pe11→pe12→pe22]","BMC","8","Open"),
    ("SYS-A07","systolic","p_valid_21_deasserts_after_start2","VP","P2",
     "$fell(sys_start_2) |=> ##[0:1] $fell(sys_valid_out_21)","BMC","8","Open"),
    ("SYS-A08","systolic","p_col_size_1_disables_col2_valid","ME","P1",
     "(ub_rd_col_size_valid_in && col_size==1) |=> !sys_valid_out_22","BMC","8","Open"),
    ("SYS-A09","systolic","p_col_size_2_both_outputs_reachable","VP","P2",
     "(col_size==2 && start_1 && start_2) |=> ##[1:6] (valid_21 && valid_22)","BMC","8","Open"),
    ("SYS-A10","systolic","p_rst_sets_pe_enabled_default","RST","P1",
     "rst |=> pe_enabled==2'b11","Unbounded","–","Open"),
    ("SYS-A11","systolic","p_pe_enabled_mask_col_size_1","SD","P2",
     "(col_size_valid && col_size==1) |=> pe_enabled==2'b01","BMC","4","Open"),
    ("SYS-A12","systolic","p_pe_enabled_mask_col_size_2","SD","P2",
     "(col_size_valid && col_size==2) |=> pe_enabled==2'b11","BMC","4","Open"),
    ("SYS-A13","systolic","p_col1_weight_load_no_col2_valid","ME","P1",
     "(sys_accept_w_1 && !sys_accept_w_2 && !sys_start_1 && !sys_start_2) |=> !sys_valid_out_22  [col weight-load independence]","BMC","4","Open"),

    # ── bias_child ────────────────────────────────────────────────────────────
    ("BC-A01","bias_child","p_rst_clears_valid","RST","P1",
     "rst |=> !bias_Z_valid_out","Unbounded","–","Open"),
    ("BC-A02","bias_child","p_rst_clears_data","RST","P1",
     "rst |=> bias_z_data_out==0","Unbounded","–","Open"),
    ("BC-A03","bias_child","p_valid_out_mirrors_valid_in","VP","P1",
     "1'b1 |=> bias_Z_valid_out==$past(bias_sys_valid_in)","BMC","4","Open"),
    ("BC-A04","bias_child","p_data_zero_when_invalid","DP","P1",
     "!bias_sys_valid_in |=> bias_z_data_out==0","BMC","4","Open"),
    ("BC-A05","bias_child","p_data_nonzero_when_both_inputs_nonzero","FA","P2",
     "(valid_in && data_in!=0 && scalar!=0) |=> bias_z_data_out!=0","BMC","4","Open"),
    ("BC-A06","bias_child","p_rst_clears_overflow","RST","P1",
     "rst |=> !bias_overflow_out  [BUG-OVF-1: sticky flag cleared on reset]","Unbounded","\u2013","Open"),
    ("BC-A07","bias_child","p_overflow_is_sticky","FA","P1",
     "bias_overflow_out |=> bias_overflow_out  [BUG-OVF-1: once set stays set until rst]","Unbounded","\u2013","Open"),

    # ── leaky_relu_child ──────────────────────────────────────────────────────
    ("LR-A01","leaky_relu_child","p_rst_clears_outputs","RST","P1",
     "rst |=> (!lr_valid_out && lr_data_out==0)","Unbounded","–","Open"),
    ("LR-A02","leaky_relu_child","p_valid_out_mirrors_input","VP","P1",
     "1'b1 |=> lr_valid_out==$past(lr_valid_in)  [gated: 0 when invalid]","BMC","4","Open"),
    ("LR-A03","leaky_relu_child","p_data_zero_when_invalid","DP","P1",
     "!lr_valid_in |=> lr_data_out==0","BMC","4","Open"),
    ("LR-A04","leaky_relu_child","p_positive_input_passthrough","FA","P1",
     "(lr_valid_in && lr_data_in>=0) |=> lr_data_out==$past(lr_data_in)","BMC","4","Open"),
    ("LR-A05","leaky_relu_child","p_negative_input_scaled","FA","P1",
     "(lr_valid_in && lr_data_in[15] && leak_factor!=1.0) |=> lr_data_out!=$past(lr_data_in)","BMC","4","Open"),
    ("LR-A06","leaky_relu_child","p_sign_preserved_for_positive","FA","P2",
     "(lr_valid_in && !lr_data_in[15]) |=> !lr_data_out[15]","BMC","4","Open"),
    ("LR-A07","leaky_relu_child","p_zero_boundary","FA","P2",
     "(lr_valid_in && lr_data_in==0) |=> lr_data_out==0","BMC","4","Open"),
    ("LR-A08","leaky_relu_child","p_rst_clears_overflow","RST","P1",
     "rst |=> !lr_overflow_out  [BUG-OVF-1: sticky flag cleared on reset]","Unbounded","\u2013","Open"),
    ("LR-A09","leaky_relu_child","p_overflow_is_sticky","FA","P1",
     "lr_overflow_out |=> lr_overflow_out  [BUG-OVF-1: once set stays set until rst]","Unbounded","\u2013","Open"),

    # ── leaky_relu_derivative_child ───────────────────────────────────────────
    ("LRD-A01","lrd_child","p_rst_clears_outputs","RST","P1",
     "rst |=> (!lr_d_valid_out && lr_d_data_out==0)","Unbounded","–","Open"),
    ("LRD-A02","lrd_child","p_valid_out_mirrors_valid_in","VP","P1",
     "1'b1 |=> lr_d_valid_out==$past(lr_d_valid_in)  [unconditional]","BMC","4","Open"),
    ("LRD-A03","lrd_child","p_data_zero_when_invalid","DP","P1",
     "!lr_d_valid_in |=> lr_d_data_out==0","BMC","4","Open"),
    ("LRD-A04","lrd_child","p_positive_H_passes_gradient_through","FA","P1",
     "(valid_in && !H[15]) |=> data_out==$past(data_in)","BMC","4","Open"),
    ("LRD-A05","lrd_child","p_negative_H_scales_gradient","FA","P1",
     "(valid_in && H[15] && leak!=1.0) |=> data_out!=$past(data_in)","BMC","4","Open"),
    ("LRD-A06","lrd_child","p_zero_H_passes_gradient_through","FA","P2",
     "(valid_in && H==0) |=> data_out==$past(data_in)","BMC","4","Open"),
    ("LRD-A07","lrd_child","p_rst_clears_overflow","RST","P1",
     "rst |=> !lr_d_overflow_out  [BUG-OVF-1: sticky flag cleared on reset]","Unbounded","\u2013","Open"),
    ("LRD-A08","lrd_child","p_overflow_is_sticky","FA","P1",
     "lr_d_overflow_out |=> lr_d_overflow_out  [BUG-OVF-1: once set stays set until rst]","Unbounded","\u2013","Open"),

    # ── loss_child ────────────────────────────────────────────────────────────
    ("LC-A01","loss_child","p_rst_clears_gradient","RST","P1",
     "rst |=> gradient_out==0","Unbounded","–","Open"),
    ("LC-A02","loss_child","p_rst_clears_valid","RST","P1",
     "rst |=> !valid_out","Unbounded","–","Open"),
    ("LC-A03","loss_child","p_valid_out_registered","VP","P1",
     "1'b1 |=> valid_out==$past(valid_in)","BMC","4","Open"),
    ("LC-A04","loss_child","p_data_zero_when_invalid","DP","P1",
     "!valid_in |=> gradient_out==0  [BUG-LC fix: gradient output cleared to 0 when input invalid]","BMC","4","Open"),
    ("LC-A05","loss_child","p_positive_gradient_when_H_gt_Y","FA","P2",
     "(valid_in && H>Y) |=> !gradient_out[15]","BMC","4","Open"),
    ("LC-A06","loss_child","p_negative_gradient_when_H_lt_Y","FA","P2",
     "(valid_in && H<Y) |=> gradient_out[15]","BMC","4","Open"),
    ("LC-A07","loss_child","p_zero_gradient_when_H_eq_Y","FA","P2",
     "(valid_in && H==Y) |=> gradient_out==0","BMC","4","Open"),
    ("LC-A08","loss_child","p_rst_clears_overflow","RST","P1",
     "rst |=> !loss_overflow_out  [BUG-OVF-1: sticky flag cleared on reset]","Unbounded","\u2013","Open"),
    ("LC-A09","loss_child","p_overflow_is_sticky","FA","P1",
     "loss_overflow_out |=> loss_overflow_out  [BUG-OVF-1: once set stays set until rst]","Unbounded","\u2013","Open"),

    # ── gradient_descent ──────────────────────────────────────────────────────
    ("GD-A01","gradient_descent","p_rst_clears_output","RST","P1",
     "rst |=> value_updated_out==0","Unbounded","–","Open"),
    ("GD-A02","gradient_descent","p_rst_clears_done","RST","P1",
     "rst |=> !grad_descent_done_out","Unbounded","–","Open"),
    ("GD-A03","gradient_descent","p_done_one_cycle_delay","VP","P1",
     "1'b1 |=> done==$past(grad_descent_valid_in)","BMC","6","Open"),
    ("GD-A04","gradient_descent","p_output_stable_when_not_valid","DP","P1",
     "!valid_in |=> value_updated_out==$past(value_updated_out)  [HOLD, not clear]","BMC","6","Open"),
    ("GD-A05","gradient_descent","p_weight_mode_produces_output_when_valid","FA","P1",
     "(valid_in && weight_mode && inputs!=0) |=> out!=0","BMC","6","Open"),
    ("GD-A06","gradient_descent","p_done_implies_valid_was_set","VP","P1",
     "done |-> $past(valid_in)","BMC","6","Open"),
    ("GD-A07","gradient_descent","p_not_done_implies_valid_was_clear","VP","P1",
     "!done |-> !$past(valid_in)","BMC","6","Open"),
    ("GD-A08","gradient_descent","p_weight_mode_descent_direction","FA","P2",
     "(valid_in && weight_mode && pos_grad && pos_lr) |=> $signed(out)<=$signed($past(old))","BMC","6","Open"),
    ("GD-A09","gradient_descent","p_rst_clears_overflow","RST","P1",
     "rst |=> !grad_overflow_out","Unbounded","–","Open"),
    ("GD-A10","gradient_descent","p_overflow_is_sticky","FA","P1",
     "grad_overflow_out |=> grad_overflow_out  [sticky until rst]","BMC","4","Open"),

    # ── control_unit ──────────────────────────────────────────────────────────
    ("CU-A01","control_unit","p_sys_switch_bit","SD","P1",
     "sys_switch_in === instruction[0]","Comb","–","Open"),
    ("CU-A02","control_unit","p_ub_rd_start_bit","SD","P1",
     "ub_rd_start_in === instruction[1]","Comb","–","Open"),
    ("CU-A03","control_unit","p_ub_rd_transpose_bit","SD","P1",
     "ub_rd_transpose === instruction[2]","Comb","–","Open"),
    ("CU-A04","control_unit","p_ub_wr_host_valid_1_bit","SD","P1",
     "ub_wr_host_valid_in_1 === instruction[3]","Comb","–","Open"),
    ("CU-A05","control_unit","p_ub_wr_host_valid_2_bit","SD","P1",
     "ub_wr_host_valid_in_2 === instruction[4]","Comb","–","Open"),
    ("CU-A06","control_unit","p_ub_rd_col_size_field","SD","P1",
     "ub_rd_col_size === instruction[20:5]","Comb","–","Open"),
    ("CU-A07","control_unit","p_ub_rd_row_size_field","SD","P1",
     "ub_rd_row_size === instruction[36:21]","Comb","–","Open"),
    ("CU-A08","control_unit","p_ub_rd_addr_field","SD","P1",
     "ub_rd_addr_in === instruction[52:37]","Comb","–","Open"),
    ("CU-A09","control_unit","p_ub_ptr_select_field","SD","P1",
     "ub_ptr_select === instruction[61:53]","Comb","–","Open"),
    ("CU-A10","control_unit","p_host_data_1_field","SD","P1",
     "ub_wr_host_data_in_1 === instruction[77:62]","Comb","–","Open"),
    ("CU-A11","control_unit","p_host_data_2_field","SD","P1",
     "ub_wr_host_data_in_2 === instruction[93:78]","Comb","–","Open"),
    ("CU-A12","control_unit","p_vpu_data_pathway_field","SD","P1",
     "vpu_data_pathway === instruction[97:94]","Comb","–","Open"),
    ("CU-A13","control_unit","p_inv_batch_size_field","SD","P1",
     "inv_batch_size_times_two_in === instruction[113:98]","Comb","–","Open"),
    ("CU-A14","control_unit","p_vpu_leak_factor_field","SD","P1",
     "vpu_leak_factor_in === instruction[129:114]","Comb","–","Open"),
    ("CU-A15","control_unit","p_bit_field_no_overlap","SD","P1",
     "All 14 field assignments cover bits [129:0] with no gap / no overlap (structural check)","Comb","–","Open"),

    # ── vpu ──────────────────────────────────────────────────────────────────
    ("VPU-A01","vpu","p_rst_clears_valid_out","RST","P1",
     "rst |=> (!vpu_valid_out_1 && !vpu_valid_out_2)","Unbounded","–","Open"),
    ("VPU-A02","vpu","p_rst_clears_data_out","RST","P1",
     "rst |=> (vpu_data_out_1==0 && vpu_data_out_2==0)","Unbounded","–","Open"),
    ("VPU-A03","vpu","p_zero_pathway_reg_valid","VP","P1",
     "pathway==0000 |=> (valid_out_1==$past(valid_in_1) && valid_out_2==$past(valid_in_2))","BMC","4","Open"),
    ("VPU-A04","vpu","p_zero_pathway_reg_data","DP","P1",
     "pathway==0000 |=> (data_out_1==$past(data_in_1) && data_out_2==$past(data_in_2))","BMC","4","Open"),
    ("VPU-A05","vpu","p_forward_path_3cy_latency","VP","P1",
     "(pathway==1100 && valid_in_1) |=> ##2 vpu_valid_out_1  [3 cycles: bias+lr+outreg]","BMC","8","Open"),
    ("VPU-A06","vpu","p_backward_path_2cy_latency","VP","P1",
     "(pathway==0001 && valid_in_1) |=> ##1 vpu_valid_out_1  [2 cycles: lrd+outreg]","BMC","6","Open"),
    ("VPU-A07","vpu","p_transition_path_5cy_latency","VP","P1",
     "(pathway==1111 && valid_in_1) |=> ##4 vpu_valid_out_1  [5 cycles: bias+lr+loss+lrd+outreg]","BMC","10","Open"),
    ("VPU-A08","vpu","p_no_output_without_input_reg","VP","P2",
     "(pathway==0000 && !valid_in_1) |=> !vpu_valid_out_1","BMC","4","Open"),
    ("VPU-A09","vpu","p_both_columns_valid_together","VP","P2",
     "(valid_in_1 && valid_in_2 && pathway!=0000) |=> valid_out_1==valid_out_2","BMC","8","Open"),
    ("VPU-A10","vpu","p_valid_out_deasserts_after_in_drops_fwd","VP","P2",
     "(pathway==1100 && $fell(valid_in_1)) |=> ##[0:3] $fell(vpu_valid_out_1)","BMC","8","Open"),
    ("VPU-A11","vpu","p_rst_clears_last_H_cache","RST","P2",
     "rst |=> (last_H_data_1_out==0 && last_H_data_2_out==0)","Unbounded","–","Open"),
    ("VPU-A12","vpu","p_last_H_clears_when_loss_inactive","DP","P2",
     "!pathway[1] |=> last_H_data_1_out==0 && last_H_data_2_out==0","BMC","6","Open"),
    ("VPU-A13","vpu","p_last_H_registers_when_loss_active","FA","P2",
     "(pathway[1] && valid_in_1) |=> last_H_data_1_out!=0 || last_H_data_2_out!=0","BMC","6","Open"),

    # ── unified_buffer ────────────────────────────────────────────────────────
    ("UB-A01","unified_buffer","p_rst_clears_wr_ptr","RST","P1",
     "rst |=> wr_ptr==0","Unbounded","–","Open"),
    ("UB-A02","unified_buffer","p_rst_deasserts_col_size_valid","RST","P1",
     "rst |=> !ub_rd_col_size_valid_out","Unbounded","–","Open"),
    ("UB-A03","unified_buffer","p_rst_clears_input_valid","RST","P1",
     "rst |=> (!ub_rd_input_valid_out_0 && !ub_rd_input_valid_out_1)","Unbounded","–","Open"),
    ("UB-A04","unified_buffer","p_rst_clears_weight_valid","RST","P1",
     "rst |=> (!ub_rd_weight_valid_out_0 && !ub_rd_weight_valid_out_1)","Unbounded","–","Open"),
    ("UB-A05","unified_buffer","p_col_size_valid_reg_decode","VP","P1",
     "1'b1 |=> col_size_valid==$past(rd_start && ptr_select==9'd1)","BMC","8","Open"),
    ("UB-A06","unified_buffer","p_wr_ptr_increments_on_vpu_write","DP","P2",
     "(wr_valid[0] && wr_valid[1]) |=> wr_ptr==$past(wr_ptr)+2  [dual-channel write, +2]","BMC","8","Open"),
    ("UB-A07a","unified_buffer","p_rd_input_ptr_in_range","SD","P2",
     "rd_input_ptr < UNIFIED_BUFFER_WIDTH","BMC","32","Open"),
    ("UB-A07b","unified_buffer","p_rd_weight_ptr_in_range","SD","P2",
     "rd_weight_ptr < UNIFIED_BUFFER_WIDTH","BMC","32","Open"),
    ("UB-A07c","unified_buffer","p_wr_ptr_in_range","SD","P2",
     "wr_ptr < UNIFIED_BUFFER_WIDTH","BMC","32","Open"),
    ("UB-A08a","unified_buffer","p_no_vpu_host_write_collision_ch0","ME","P1",
     "!(ub_wr_valid_in[0] && ub_wr_host_valid_in[0])","BMC","8","Open"),
    ("UB-A08b","unified_buffer","p_no_vpu_host_write_collision_ch1","ME","P1",
     "!(ub_wr_valid_in[1] && ub_wr_host_valid_in[1])","BMC","8","Open"),
    ("UB-A09a","unified_buffer","p_col_size_out_correct_non_transpose","DP","P2",
     "(col_size_valid && !rd_transpose) |-> col_size_out==ub_rd_col_size","BMC","8","Open"),
    ("UB-A09b","unified_buffer","p_col_size_out_correct_transpose","DP","P2",
     "(col_size_valid && rd_transpose) |-> col_size_out==ub_rd_row_size","BMC","8","Open"),
    ("UB-A10","unified_buffer","p_col_size_out_zero_when_not_valid","DP","P2",
     "!col_size_valid |-> col_size_out==0","BMC","8","Open"),
    ("UB-A11","unified_buffer","p_rst_clears_rd_bias_ptr","RST","P1",
     "rst |=> rd_bias_ptr==0","Unbounded","–","Open"),
    ("UB-A12","unified_buffer","p_rst_clears_rd_Y_ptr","RST","P1",
     "rst |=> rd_Y_ptr==0","Unbounded","–","Open"),
    ("UB-A13","unified_buffer","p_rst_clears_rd_H_ptr","RST","P1",
     "rst |=> rd_H_ptr==0","Unbounded","–","Open"),
    ("UB-A14","unified_buffer","p_rst_clears_rd_grad_bias_ptr","RST","P1",
     "rst |=> rd_grad_bias_ptr==0","Unbounded","–","Open"),
    ("UB-A15","unified_buffer","p_rst_clears_rd_grad_weight_ptr","RST","P1",
     "rst |=> rd_grad_weight_ptr==0","Unbounded","–","Open"),
    ("UB-A16","unified_buffer","p_rst_clears_grad_descent_ptr","RST","P1",
     "rst |=> grad_descent_ptr==0","Unbounded","–","Open"),
    ("UB-A17","unified_buffer","p_grad_descent_ptr_max_advance","DP","P1",
     "grad_descent_ptr <= $past(grad_descent_ptr)+2  [at most 2 GD instances fire per cycle]","BMC","8","Open"),
    ("UB-A18","unified_buffer","p_grad_descent_ptr_monotonic","DP","P1",
     "grad_descent_ptr >= $past(grad_descent_ptr)  [no decrement without reset]","BMC","8","Open"),
]

# Category legend
CAT_LEGEND = [
    ("RST", "Reset / clear behaviour"),
    ("VP",  "Valid-protocol timing"),
    ("DP",  "Datapath correctness"),
    ("FA",  "Functional arithmetic"),
    ("SD",  "Structural / decode"),
    ("ME",  "Mutual-exclusion / protocol"),
]

# Cover properties per module
COVERS = [
    ("pe",           "PE_C1",   "MAC active (valid_in=1, pe_enabled=1)"),
    ("pe",           "PE_C2",   "Weight load → switch across two cycles"),
    ("pe",           "PE_C3",   "PE disabled mid-compute"),
    ("systolic",     "S_C1",    "Both columns active simultaneously"),
    ("systolic",     "S_C2",    "sys_switch_in asserted during computation"),
    ("systolic",     "S_C3",    "col_size_valid issued with col_size=1"),
    ("systolic",     "S_C4",    "Column-1-only weight load"),
    ("systolic",     "S_C5",    "Column-2-only weight load"),
    ("systolic",     "S_C6",    "Both rows started together"),
    ("systolic",     "S_C7",    "Row-1-only start (column-2 computation)"),
    ("bias_child",   "BC_C1",   "Positive + positive inputs"),
    ("bias_child",   "BC_C2",   "Negative + positive (sign change)"),
    ("bias_child",   "BC_C3",   "valid deasserted mid-burst"),
    ("lrc_child",    "LR_C1",   "Positive input passthrough"),
    ("lrc_child",    "LR_C2",   "Negative input scaled"),
    ("lrc_child",    "LR_C3",   "Zero boundary"),
    ("lrc_child",    "LR_C4",   "valid deasserted"),
    ("lrd_child",    "LRD_C1",  "H > 0: gradient passthrough"),
    ("lrd_child",    "LRD_C2",  "H < 0: gradient scaled"),
    ("lrd_child",    "LRD_C3",  "H = 0 boundary"),
    ("lrd_child",    "LRD_C4",  "valid deasserted"),
    ("loss_child",   "LC_C1",   "H > Y: positive gradient"),
    ("loss_child",   "LC_C2",   "H < Y: negative gradient"),
    ("loss_child",   "LC_C3",   "H = Y: zero gradient"),
    ("loss_child",   "LC_C4",   "valid deasserted"),
    ("gradient_desc","GD_C1",   "Weight mode (grad_bias_or_weight=1)"),
    ("gradient_desc","GD_C2",   "Bias mode (grad_bias_or_weight=0)"),
    ("gradient_desc","GD_C3",   "Bias mode accumulation (done cascades)"),
    ("gradient_desc","GD_C4",   "valid_in deasserted after a run ($fell(valid_in))"),
    ("control_unit", "CU_C1",   "Forward pass pathway (1100)"),
    ("control_unit", "CU_C2",   "Transition pathway (1111)"),
    ("control_unit", "CU_C3",   "Backward pass pathway (0001)"),
    ("control_unit", "CU_C4",   "Passthrough pathway (0000)"),
    ("control_unit", "CU_C5",   "sys_switch_in asserted"),
    ("control_unit", "CU_C6",   "ub_rd_transpose asserted"),
    ("control_unit", "CU_C7",   "Transposed read (start && transpose)"),
    ("vpu",          "VPU_C1",  "Forward path completes (valid_out_1 seen)"),
    ("vpu",          "VPU_C2",  "Transition path completes"),
    ("vpu",          "VPU_C3",  "Backward path completes"),
    ("vpu",          "VPU_C4",  "Zero pathway passthrough"),
    ("vpu",          "VPU_C5",  "Both channels active simultaneously"),
    ("unified_buffer","UB_COV1","Full input read burst (row=2, col=2)"),
    ("unified_buffer","UB_COV2","Full weight read burst"),
    ("unified_buffer","UB_COV3","Host write followed by VPU write-back"),
    ("unified_buffer","UB_COV4","Transpose read exercised"),
    ("unified_buffer","UB_COV5","Non-transpose read exercised"),
    ("unified_buffer","UB_COV6","col_size_valid output asserted"),
    ("unified_buffer","UB_COV7","wr_ptr reaches ≥4 (2 full dual-channel write-back cycles)"),
    ("unified_buffer","UB_COV8","Both VPU channels write simultaneously (wr_valid[0] && wr_valid[1])"),
    ("unified_buffer","UB_COV9","grad_descent_ptr advances past zero (write-back chain exercised)"),
]

# Assumptions per module
ASSUMPTIONS = [
    ("pe",           "PE_ASM_01", "pe_enabled is held constant during a computation burst"),
    ("pe",           "PE_ASM_02", "psum_in=0 for the top row (row-1 PEs, no upstream partial sum)"),
    ("systolic",     "SYS_ASM_01","ub_rd_col_size_in ∈ {1, 2} when valid"),
    ("systolic",     "SYS_ASM_02","sys_start_1 deasserted ≥1 cycle between batches"),
    ("systolic",     "SYS_ASM_02b","sys_start_2 deasserted ≥1 cycle between batches"),
    ("systolic",     "SYS_ASM_03","sys_accept_w_1 and sys_accept_w_2 never simultaneously asserted"),
    ("gradient_desc","GD_ASM_04","Learning rate always positive (lr_in[15]=0, lr_in≠0)"),
    ("vpu",          "VPU_ASM_01","vpu_data_pathway ∈ {0000, 1100, 1111, 0001}"),
    ("vpu",          "VPU_ASM_02","pathway stable during a burst (no mid-burst change)"),
    ("vpu",          "VPU_ASM_04","Bias scalars = 0 when bias stage inactive (pathway[3]=0)"),
    ("unified_buffer","UB_ASM_01","ub_ptr_select < 8 (valid pointer range)"),
    ("unified_buffer","UB_ASM_02a","Host and VPU writes mutually exclusive on channel 0"),
    ("unified_buffer","UB_ASM_02b","Host and VPU writes mutually exclusive on channel 1"),
    ("unified_buffer","UB_ASM_03a","rd_row_size ≥1 and ≤ SYSTOLIC_ARRAY_WIDTH"),
    ("unified_buffer","UB_ASM_03b","rd_col_size ≥1 and ≤ SYSTOLIC_ARRAY_WIDTH"),
    ("unified_buffer","UB_ASM_04","learning_rate always positive"),
    ("unified_buffer","UB_ASM_05","ub_rd_start_in=0 on first cycle after reset deasserts"),
]

# ════════════════════════════════════════════════════════════════════════════
#  BUILD WORKBOOK
# ════════════════════════════════════════════════════════════════════════════

wb = openpyxl.Workbook()

# ── SHEET 1 : Master Assertion Catalog ──────────────────────────────────────
ws_master = wb.active
ws_master.title = "Master Assertions"
ws_master.sheet_view.showGridLines = False

COLS_MASTER = ["Assert ID", "Module", "SVA Property Name", "Category",
               "Priority", "Property / Description", "Proof Type",
               "Bound (k)", "Status"]

# freeze panes
ws_master.freeze_panes = "A3"

# title row
hdr_cell(ws_master, 1, 1, "tiny-tpu Formal Verification — Master Assertion Catalog", merge_end=9)

# column headers
for ci, h in enumerate(COLS_MASTER, 1):
    hdr_cell(ws_master, 2, ci, h)

# Column widths
widths = [10, 18, 38, 10, 9, 60, 12, 9, 9]
for ci, w in enumerate(widths, 1):
    ws_master.column_dimensions[get_column_letter(ci)].width = w

# Data rows
prev_mod = None
row = 3
for i, a in enumerate(ASSERTIONS):
    alt = (i % 2 == 0)
    mod = a[1]
    if mod != prev_mod:
        # section separator
        sub_cell(ws_master, row, 1, f"── {mod.upper()} ──", merge_end=9)
        ws_master.row_dimensions[row].height = 18
        row += 1
        prev_mod = mod
    for ci, val in enumerate(a, 1):
        if ci == 6:   # property column – monospace
            body_cell(ws_master, row, ci, val, alt=alt, mono=True)
        elif ci in (1, 4, 5, 7, 8):   # ID/cat/prio/proof/bound – centred
            body_cell(ws_master, row, ci, val, alt=alt, center=True)
        elif ci == 9:   # status
            status_cell(ws_master, row, ci, val)
            if alt:
                pass  # status_cell sets its own fill
        else:
            body_cell(ws_master, row, ci, val, alt=alt)
    ws_master.row_dimensions[row].height = 28
    row += 1

# ── SHEET 2 : Cover Properties ───────────────────────────────────────────────
ws_cov = wb.create_sheet("Cover Properties")
ws_cov.sheet_view.showGridLines = False
ws_cov.freeze_panes = "A3"

hdr_cell(ws_cov, 1, 1, "tiny-tpu Formal Verification — Cover Properties", merge_end=3)
for ci, h in enumerate(["Module", "Cover ID", "Description"], 1):
    hdr_cell(ws_cov, 2, ci, h)
ws_cov.column_dimensions["A"].width = 18
ws_cov.column_dimensions["B"].width = 14
ws_cov.column_dimensions["C"].width = 60

prev_mod = None
row = 3
for i, (mod, cid, desc) in enumerate(COVERS):
    alt = (i % 2 == 0)
    if mod != prev_mod:
        sub_cell(ws_cov, row, 1, f"── {mod.upper()} ──", merge_end=3)
        row += 1
        prev_mod = mod
    body_cell(ws_cov, row, 1, mod, alt=alt)
    body_cell(ws_cov, row, 2, cid, alt=alt, center=True)
    body_cell(ws_cov, row, 3, desc, alt=alt)
    ws_cov.row_dimensions[row].height = 22
    row += 1

# ── SHEET 3 : Assumptions ────────────────────────────────────────────────────
ws_asm = wb.create_sheet("Assumptions")
ws_asm.sheet_view.showGridLines = False
ws_asm.freeze_panes = "A3"

hdr_cell(ws_asm, 1, 1, "tiny-tpu Formal Verification — Formal Assumptions (Constraints)", merge_end=3)
for ci, h in enumerate(["Module", "Assume ID", "Description"], 1):
    hdr_cell(ws_asm, 2, ci, h)
ws_asm.column_dimensions["A"].width = 18
ws_asm.column_dimensions["B"].width = 16
ws_asm.column_dimensions["C"].width = 70

prev_mod = None
row = 3
for i, (mod, aid, desc) in enumerate(ASSUMPTIONS):
    alt = (i % 2 == 0)
    if mod != prev_mod:
        sub_cell(ws_asm, row, 1, f"── {mod.upper()} ──", merge_end=3)
        row += 1
        prev_mod = mod
    body_cell(ws_asm, row, 1, mod, alt=alt)
    body_cell(ws_asm, row, 2, aid, alt=alt, center=True)
    body_cell(ws_asm, row, 3, desc, alt=alt)
    ws_asm.row_dimensions[row].height = 22
    row += 1

# ── SHEET 4 : Module Summary ─────────────────────────────────────────────────
ws_sum = wb.create_sheet("Module Summary")
ws_sum.sheet_view.showGridLines = False
ws_sum.freeze_panes = "A3"

hdr_cell(ws_sum, 1, 1, "tiny-tpu Formal Verification — Module Summary", merge_end=7)
for ci, h in enumerate(["Module", "RTL File", "SVA File",
                         "Assertions", "Covers", "Proof Bound k",
                         "Verification Approach"], 1):
    hdr_cell(ws_sum, 2, ci, h)

SUMMARY = [
    ("pe",            "src/pe.sv",              "sva/pe_assertions.sv",              23, 3,  6,  "BMC"),
    ("systolic",      "src/systolic.sv",         "sva/systolic_assertions.sv",        13, 7,  8,  "BMC"),
    ("bias_child",    "src/bias_child.sv",       "sva/bias_child_assertions.sv",       7, 3,  4,  "BMC"),
    ("leaky_relu_child","src/leaky_relu_child.sv","sva/leaky_relu_child_assertions.sv", 9, 4,  4,  "BMC"),
    ("lrd_child",     "src/leaky_relu_derivative_child.sv","sva/leaky_relu_derivative_child_assertions.sv",8,4,4,"BMC"),
    ("loss_child",    "src/loss_child.sv",       "sva/loss_child_assertions.sv",       9, 4,  4,  "BMC"),
    ("gradient_descent","src/gradient_descent.sv","sva/gradient_descent_assertions.sv",10,4,  6,  "BMC"),
    ("control_unit",  "src/control_unit.sv",     "sva/control_unit_assertions.sv",    15, 7,  0,  "Comb (k=0)"),
    ("vpu",           "src/vpu.sv",              "sva/vpu_assertions.sv",             13, 5,  10, "BMC"),
    ("unified_buffer","src/unified_buffer.sv",   "sva/unified_buffer_assertions.sv",  22, 9,  32, "BMC+Induction"),
]

widths_sum = [18, 30, 44, 12, 9, 13, 22]
for ci, w in enumerate(widths_sum, 1):
    ws_sum.column_dimensions[get_column_letter(ci)].width = w

for ri, (mod, rtl, sva, na, nc, k, approach) in enumerate(SUMMARY, 3):
    alt = (ri % 2 == 0)
    body_cell(ws_sum, ri, 1, mod, alt=alt)
    body_cell(ws_sum, ri, 2, rtl, alt=alt, mono=True)
    body_cell(ws_sum, ri, 3, sva, alt=alt, mono=True)
    body_cell(ws_sum, ri, 4, na,  alt=alt, center=True)
    body_cell(ws_sum, ri, 5, nc,  alt=alt, center=True)
    body_cell(ws_sum, ri, 6, k,   alt=alt, center=True)
    body_cell(ws_sum, ri, 7, approach, alt=alt)
    ws_sum.row_dimensions[ri].height = 22

# totals row
tr = len(SUMMARY) + 3
c = ws_sum.cell(row=tr, column=1, value="TOTAL")
c.font = Font(name="Calibri", bold=True, size=10)
c.fill = SUB_FILL; c.font = SUB_FONT; c.alignment = CTR; c.border = THIN_BORDER
total_a = sum(r[3] for r in SUMMARY)
total_c = sum(r[4] for r in SUMMARY)
for ci, val in zip([4, 5], [total_a, total_c]):
    cx = ws_sum.cell(row=tr, column=ci, value=val)
    cx.font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    cx.fill = SUB_FILL; cx.alignment = CTR; cx.border = THIN_BORDER
for ci in [2, 3, 6, 7]:
    cx = ws_sum.cell(row=tr, column=ci, value="")
    cx.fill = SUB_FILL; cx.border = THIN_BORDER

# ── SHEET 5 : Category Legend ────────────────────────────────────────────────
ws_leg = wb.create_sheet("Legend")
ws_leg.sheet_view.showGridLines = False
hdr_cell(ws_leg, 1, 1, "Category Codes", merge_end=2)
hdr_cell(ws_leg, 2, 1, "Code")
hdr_cell(ws_leg, 2, 2, "Meaning")
ws_leg.column_dimensions["A"].width = 10
ws_leg.column_dimensions["B"].width = 40
for ri, (code, meaning) in enumerate(CAT_LEGEND, 3):
    alt = (ri % 2 == 0)
    body_cell(ws_leg, ri, 1, code, alt=alt, center=True)
    body_cell(ws_leg, ri, 2, meaning, alt=alt)

# Status legend
hdr_cell(ws_leg, 11, 1, "Status Colours", merge_end=2)
sl = [("Open", OPEN_FILL, "Not yet run / in progress"),
      ("Pass/Proven", PASS_FILL, "Formally proven or simulation passed"),
      ("Fail/CEX", FAIL_FILL, "Counterexample found or test failed")]
for ri, (s, fill, desc) in enumerate(sl, 12):
    c1 = ws_leg.cell(row=ri, column=1, value=s)
    c1.fill = fill; c1.font = Font(name="Calibri", bold=True, size=10)
    c1.alignment = CTR; c1.border = THIN_BORDER
    c2 = ws_leg.cell(row=ri, column=2, value=desc)
    c2.fill = WHITE_FILL; c2.font = BODY_FONT
    c2.alignment = WRAP; c2.border = THIN_BORDER

# ── Save ─────────────────────────────────────────────────────────────────────
out_path = os.path.join(os.path.dirname(__file__), "tiny_tpu_FV_Plan.xlsx")
wb.save(out_path)
print(f"Saved: {out_path}")
print(f"Total assertions : {total_a}")
print(f"Total covers     : {total_c}")
